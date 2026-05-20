import csv
import hashlib
import logging
import unicodedata
import hmac
import uuid
from datetime import datetime, timedelta, date
from functools import lru_cache
from io import BytesIO, StringIO

import requests
from django.conf import settings
from django.db import connection
from django.contrib.auth.hashers import make_password, check_password, identify_hasher
from openpyxl import load_workbook, Workbook

from .models import Conta, Produto

logger = logging.getLogger(__name__)


def hash_senha_antiga(senha: str) -> str:
    """Hash antigo usado no executável. Mantido para compatibilidade com contas já criadas."""
    return hashlib.sha256((senha or '').encode('utf-8')).hexdigest()


def hash_senha(senha: str) -> str:
    """Hash seguro para novas contas web."""
    return make_password(senha or '')


def gerar_token() -> str:
    return str(uuid.uuid4())


def normalizar_email(valor: str) -> str:
    valor = (valor or '').strip().lower()
    if valor and '@' not in valor:
        valor = f'{valor}@app.com'
    return valor


def verificar_senha(senha_digitada: str, senha_hash_salva: str) -> bool:
    senha_hash_salva = senha_hash_salva or ''
    try:
        identify_hasher(senha_hash_salva)
        return check_password(senha_digitada or '', senha_hash_salva)
    except Exception:
        # Compatibilidade com o hash SHA256 do aplicativo antigo.
        return hmac.compare_digest(hash_senha_antiga(senha_digitada), senha_hash_salva)


def get_client_ip(request) -> str:
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def calcular_dias_restantes(data_expiracao) -> int:
    if not data_expiracao:
        return 0
    if isinstance(data_expiracao, datetime):
        data_expiracao = data_expiracao.date()
    hoje = date.today()
    return max(0, (data_expiracao - hoje).days)


def conta_esta_ativa(conta: Conta) -> bool:
    return bool(conta and int(conta.ativo or 0) == 1)


def conta_trial_expirado(conta: Conta) -> bool:
    return not conta_esta_ativa(conta) and calcular_dias_restantes(conta.trial_expira_em) <= 0


def limite_produtos(conta: Conta) -> int:
    return settings.PRO_LIMITE_PRODUTOS if conta_esta_ativa(conta) else settings.TRIAL_LIMITE_PRODUTOS


def registrar_log(email: str, acao: str) -> None:
    try:
        with connection.cursor() as cursor:
            cursor.execute('INSERT INTO logs (email, acao, criado_em) VALUES (%s, %s, %s)', [email, acao, datetime.now()])
    except Exception:
        pass


def criar_conta(email: str, senha: str, ip: str):
    email = normalizar_email(email)
    if not email or not senha:
        return None, 'Preencha usuário/e-mail e senha.'
    if len(senha) < 4:
        return None, 'A senha precisa ter pelo menos 4 caracteres.'

    if Conta.objects.filter(email=email).exists():
        return None, 'Este usuário já existe.'

    total_ip = Conta.objects.filter(ip=ip, ativo=0).count() if ip else 0
    if total_ip >= settings.TRIAL_MAX_CONTAS_POR_IP:
        return None, 'Limite de contas trial atingido neste IP.'

    agora = datetime.now()
    conta = Conta(
        email=email,
        senha=hash_senha(senha),
        token=gerar_token(),
        criado_em=agora,
        trial_expira_em=agora + timedelta(days=settings.TRIAL_DIAS),
        ativo=0,
        device_id='web',
        plano='trial',
        ip=ip,
    )
    conta.save(force_insert=True)
    registrar_log(email, 'registro_web')
    return conta, None


def autenticar(email: str, senha: str):
    email = normalizar_email(email)
    try:
        conta = Conta.objects.get(email=email)
    except Conta.DoesNotExist:
        return None, 'Usuário não encontrado.'

    if not verificar_senha(senha, conta.senha):
        return None, 'Senha inválida.'

    # Mesmo com trial vencido, deixa entrar para conseguir ver painel e pagar.
    conta.token = gerar_token()
    conta.save(update_fields=['token'])
    registrar_log(email, 'login_web')
    return conta, None


def get_conta_por_email(email: str):
    if not email:
        return None
    try:
        return Conta.objects.get(email=email)
    except Conta.DoesNotExist:
        return None


def produto_status(validade: str) -> str:
    try:
        data = datetime.strptime(str(validade), '%Y-%m-%d').date()
    except Exception:
        return 'sem-data'

    hoje = date.today()
    if data < hoje:
        return 'vencido'
    if (data - hoje).days <= settings.VENCIMENTO_PROXIMO_DIAS:
        return 'proximo'
    return 'ok'


def status_label(status: str) -> str:
    return {
        'vencido': 'Vencido',
        'proximo': 'Próximo',
        'ok': 'OK',
        'sem-data': 'Sem data',
    }.get(status, status)


def formatar_data_br(validade: str) -> str:
    try:
        return datetime.strptime(str(validade), '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return validade or '-'


def produtos_do_usuario(email: str):
    return Produto.objects.filter(user_email=email).order_by('validade', 'nome')


def estatisticas(conta: Conta) -> dict:
    qs = produtos_do_usuario(conta.email)
    produtos = list(qs)
    total = len(produtos)
    vencidos = sum(1 for p in produtos if produto_status(p.validade) == 'vencido')
    proximos = sum(1 for p in produtos if produto_status(p.validade) == 'proximo')
    ok = sum(1 for p in produtos if produto_status(p.validade) == 'ok')
    limite = limite_produtos(conta)
    uso = min(100, int((total / max(1, limite)) * 100))
    expirado = conta_trial_expirado(conta)
    return {
        'total': total,
        'vencidos': vencidos,
        'proximos': proximos,
        'ok': ok,
        'trial_restante': calcular_dias_restantes(conta.trial_expira_em),
        'limite': limite,
        'uso': uso,
        'plano': 'PRO' if conta_esta_ativa(conta) else 'TRIAL',
        'ativo': conta_esta_ativa(conta),
        'trial_expirado': expirado,
        'bloqueado': total >= limite or expirado,
    }


def pode_adicionar_produto(conta: Conta):
    if conta_trial_expirado(conta):
        return False, 'Trial expirado. Atualize para o PRO para continuar cadastrando produtos.'
    total = Produto.objects.filter(user_email=conta.email).count()
    limite = limite_produtos(conta)
    if total >= limite:
        return False, f'Limite atingido: {total}/{limite} produtos.'
    return True, None


def _normalizar_texto_catalogo(valor) -> str:
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = texto.replace('_', ' ').replace('-', ' ')
    return ' '.join(texto.split())


def _normalizar_codigo_catalogo(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip().replace('\u00a0', '')
    if texto.endswith('.0') and texto[:-2].isdigit():
        texto = texto[:-2]
    return texto.strip()


@lru_cache(maxsize=1)
def carregar_catalogo_produtos():
    """Lê data/dados.xlsx uma vez e cria busca por código interno e GTIN.

    Alguns relatórios XLSX exportados pelo sistema antigo vêm com a dimensão da
    planilha marcada como A1:A1. Sem reset_dimensions(), o openpyxl enxerga só a
    coluna A e a base carrega com 0 produtos.
    """
    path = settings.DATA_EXCEL_PATH
    if not path.exists():
        return {}

    catalogo = {}
    try:
        # Não usar read_only=True aqui: este XLSX vem com dimensão interna incorreta
        # em modo streaming e o openpyxl lê apenas a coluna A.
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.active

        # Corrige planilhas exportadas com dimensão incorreta.
        try:
            ws.reset_dimensions()
        except Exception:
            pass

        header_row = None
        headers = []

        # Procura automaticamente a linha do cabeçalho. No relatório original
        # costuma ser a linha 39, mas isso deixa o leitor mais seguro.
        for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=120, values_only=True), start=1):
            normalizados = [_normalizar_texto_catalogo(c) for c in row]
            if 'produto' in normalizados and ('descricao' in normalizados or 'nome' in normalizados):
                header_row = row_number
                headers = normalizados
                break

        if header_row is None:
            return {}

        def col_idx(*nomes):
            alvos = [_normalizar_texto_catalogo(nome) for nome in nomes]
            for alvo in alvos:
                if alvo in headers:
                    return headers.index(alvo)
            return None

        idx_produto = col_idx('Produto', 'Código', 'Codigo', 'Código interno', 'Codigo interno')
        idx_descricao = col_idx('Descrição', 'Descricao', 'Nome', 'Nome do produto', 'Produto descrição')
        idx_gtin = col_idx('GTIN Principal', 'GTIN', 'EAN', 'Código de barras', 'Codigo de barras')

        if idx_descricao is None or (idx_produto is None and idx_gtin is None):
            return {}

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            codigo = row[idx_produto] if idx_produto is not None and idx_produto < len(row) else None
            nome = row[idx_descricao] if idx_descricao is not None and idx_descricao < len(row) else None
            gtin = row[idx_gtin] if idx_gtin is not None and idx_gtin < len(row) else None

            nome_s = str(nome or '').strip()
            codigo_s = _normalizar_codigo_catalogo(codigo)
            gtin_s = _normalizar_codigo_catalogo(gtin)

            if not nome_s or (not codigo_s and not gtin_s):
                continue

            info = {'codigo': codigo_s or gtin_s, 'nome': nome_s, 'gtin': gtin_s}

            if codigo_s and codigo_s.lower() != 'none':
                catalogo[codigo_s] = info
            if gtin_s and gtin_s.lower() != 'none':
                catalogo[gtin_s] = info

        try:
            wb.close()
        except Exception:
            pass
    except Exception as exc:
        logger.exception('Falha ao carregar catálogo de produtos em %s: %s', path, exc)
        return {}

    return catalogo


def buscar_catalogo(codigo: str):
    codigo = _normalizar_codigo_catalogo(codigo)
    if not codigo:
        return None
    return carregar_catalogo_produtos().get(codigo)


def limpar_validade(valor) -> str:
    """Aceita yyyy-mm-dd, dd/mm/yyyy ou data Excel e devolve yyyy-mm-dd."""
    if not valor:
        return ''
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    if isinstance(valor, date):
        return valor.strftime('%Y-%m-%d')
    texto = str(valor).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(texto, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return texto


def montar_produto_dict(produto: Produto) -> dict:
    status = produto_status(produto.validade)
    return {
        'id': produto.id,
        'codigo': produto.codigo or '',
        'nome': produto.nome or '',
        'validade_original': produto.validade or '',
        'validade': formatar_data_br(produto.validade),
        'quantidade': produto.quantidade,
        'tipo_qtd': produto.tipo_qtd or 'Un',
        'status': status,
        'status_label': status_label(status),
    }


def exportar_produtos_xlsx(conta: Conta) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produtos'
    ws.append(['ID', 'Código', 'Nome', 'Validade', 'Quantidade', 'Tipo', 'Status'])
    for produto in produtos_do_usuario(conta.email):
        item = montar_produto_dict(produto)
        ws.append([
            item['id'], item['codigo'], item['nome'], item['validade'],
            item['quantidade'], item['tipo_qtd'], item['status_label']
        ])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 48)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def exportar_produtos_csv(conta: Conta) -> str:
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['ID', 'Código', 'Nome', 'Validade', 'Quantidade', 'Tipo', 'Status'])
    for produto in produtos_do_usuario(conta.email):
        item = montar_produto_dict(produto)
        writer.writerow([
            item['id'], item['codigo'], item['nome'], item['validade'],
            item['quantidade'], item['tipo_qtd'], item['status_label']
        ])
    return output.getvalue()


def importar_produtos_de_planilha(conta: Conta, arquivo) -> tuple[int, list[str]]:
    """Importa XLSX/CSV com colunas código, nome, validade, quantidade e tipo."""
    erros = []
    criados = 0
    nome_arquivo = (getattr(arquivo, 'name', '') or '').lower()

    def normalizar_header(h):
        return str(h or '').strip().lower().replace('ó', 'o').replace('ç', 'c')

    def processar_linhas(linhas):
        nonlocal criados
        linhas = list(linhas)
        if not linhas:
            return
        headers = [normalizar_header(h) for h in linhas[0]]

        def idx(*nomes):
            for nome in nomes:
                nome = normalizar_header(nome)
                if nome in headers:
                    return headers.index(nome)
            return None

        idx_codigo = idx('codigo', 'código', 'gtin', 'codigo/gtin')
        idx_nome = idx('nome', 'descricao', 'descrição', 'produto')
        idx_validade = idx('validade', 'vencimento', 'data')
        idx_qtd = idx('quantidade', 'qtd')
        idx_tipo = idx('tipo', 'tipo_qtd', 'unidade')

        if idx_codigo is None or idx_nome is None or idx_validade is None:
            erros.append('A planilha precisa ter pelo menos: código, nome e validade.')
            return

        for numero, row in enumerate(linhas[1:], start=2):
            try:
                codigo = str(row[idx_codigo] or '').replace('.0', '').strip() if idx_codigo < len(row) else ''
                nome = str(row[idx_nome] or '').strip() if idx_nome < len(row) else ''
                validade = limpar_validade(row[idx_validade] if idx_validade < len(row) else '')
                qtd_raw = row[idx_qtd] if idx_qtd is not None and idx_qtd < len(row) else 0
                tipo = str(row[idx_tipo] or 'Un').strip() if idx_tipo is not None and idx_tipo < len(row) else 'Un'
                quantidade = int(float(qtd_raw or 0))
                datetime.strptime(validade, '%Y-%m-%d')
                if not codigo or not nome:
                    continue
                ok, erro = pode_adicionar_produto(conta)
                if not ok:
                    erros.append(erro)
                    break
                Produto.objects.create(
                    codigo=codigo,
                    nome=nome,
                    validade=validade,
                    quantidade=max(0, quantidade),
                    tipo_qtd=tipo or 'Un',
                    user_email=conta.email,
                )
                criados += 1
            except Exception:
                erros.append(f'Linha {numero}: dados inválidos.')

    if nome_arquivo.endswith('.csv'):
        conteudo = arquivo.read().decode('utf-8-sig', errors='ignore')
        sample = conteudo[:2048]
        delimiter = ';' if sample.count(';') >= sample.count(',') else ','
        processar_linhas(csv.reader(StringIO(conteudo), delimiter=delimiter))
    else:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        processar_linhas(ws.iter_rows(values_only=True))

    if criados:
        registrar_log(conta.email, f'importacao_{criados}_produtos_web')
    return criados, erros[:5]


ASAAS_BASE_URL = 'https://api.asaas.com/v3'


def asaas_headers():
    if not settings.ASAAS_API_KEY:
        return None
    return {
        'access_token': settings.ASAAS_API_KEY,
        'Content-Type': 'application/json',
    }


def criar_cliente_asaas(email: str):
    headers = asaas_headers()
    if not headers:
        return None, 'ASAAS_API_KEY não configurada.'

    try:
        response = requests.post(
            f'{ASAAS_BASE_URL}/customers',
            headers=headers,
            json={'name': email, 'email': email},
            timeout=30,
        )
        try:
            data = response.json()
        except Exception:
            data = {}
        if response.status_code >= 400 or 'errors' in data:
            consulta = requests.get(
                f'{ASAAS_BASE_URL}/customers',
                headers=headers,
                params={'email': email},
                timeout=30,
            )
            clientes = consulta.json().get('data', [])
            if clientes:
                return clientes[0].get('id'), None
            return None, 'Não foi possível criar/consultar cliente no Asaas.'
        return data.get('id'), None
    except Exception as exc:
        return None, f'Falha ao conectar no Asaas: {exc}'


def criar_pagamento_pix(email: str):
    headers = asaas_headers()
    if not headers:
        return None, 'ASAAS_API_KEY não configurada no ambiente.'

    customer_id, erro = criar_cliente_asaas(email)
    if erro:
        return None, erro

    try:
        response = requests.post(
            f'{ASAAS_BASE_URL}/payments',
            headers=headers,
            json={
                'customer': customer_id,
                'billingType': 'PIX',
                'value': settings.PIX_VALOR,
                'description': settings.PAGAMENTO_DESCRICAO,
                'externalReference': email,
            },
            timeout=30,
        )
        try:
            data = response.json()
        except Exception:
            data = {}
        if response.status_code >= 400 or 'errors' in data:
            return None, 'Erro ao criar pagamento no Asaas. Confira a ASAAS_API_KEY.'

        payment_id = data.get('id')
        qr_response = requests.get(
            f'{ASAAS_BASE_URL}/payments/{payment_id}/pixQrCode',
            headers=headers,
            timeout=30,
        )
        qr_data = qr_response.json()
        if qr_response.status_code >= 400 or 'errors' in qr_data:
            return None, 'Erro ao gerar QR Code PIX.'

        with connection.cursor() as cursor:
            cursor.execute(
                'INSERT INTO pagamentos (payment_id, email, status, criado_em) VALUES (%s, %s, %s, %s) ON CONFLICT (payment_id) DO NOTHING'
                if connection.vendor == 'postgresql'
                else 'INSERT OR IGNORE INTO pagamentos (payment_id, email, status, criado_em) VALUES (%s, %s, %s, %s)',
                [payment_id, email, data.get('status', 'PENDING'), datetime.now()],
            )

        return {
            'payment_id': payment_id,
            'qr': qr_data.get('payload'),
            'qr_base64': qr_data.get('encodedImage'),
            'valor': settings.PIX_VALOR,
        }, None
    except Exception as exc:
        return None, f'Falha ao criar PIX: {exc}'


def ativar_usuario(email: str, dias: int = 30) -> bool:
    try:
        conta = Conta.objects.get(email=normalizar_email(email))
    except Conta.DoesNotExist:
        return False
    conta.ativo = 1
    conta.plano = 'pago'
    conta.trial_expira_em = datetime.now() + timedelta(days=dias)
    conta.save(update_fields=['ativo', 'plano', 'trial_expira_em'])
    registrar_log(conta.email, 'pagamento_aprovado_web')
    return True
