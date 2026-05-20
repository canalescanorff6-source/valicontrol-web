import csv
import hashlib
import logging
import unicodedata
import hmac
import os
import re
import sqlite3
import tempfile
import uuid
import secrets
import urllib.parse
from pathlib import Path
from zipfile import ZipFile
from xml.etree.ElementTree import iterparse
from datetime import datetime, timedelta, date
from functools import lru_cache
from io import BytesIO, StringIO

import requests
from django.conf import settings
from django.db import connection
from django.contrib.auth.hashers import make_password, check_password, identify_hasher
from openpyxl import load_workbook, Workbook

from .models import Conta, Produto

logger = logging.getLogger(__name__)


def hash_senha_antiga(senha: str) -> str:
    """Hash antigo usado no executável. Mantido para compatibilidade com contas já criadas."""
    return hashlib.sha256((senha or '').encode('utf-8')).hexdigest()


def hash_senha(senha: str) -> str:
    """Hash seguro para novas contas web."""
    return make_password(senha or '')


def gerar_token() -> str:
    return str(uuid.uuid4())


def normalizar_email(valor: str) -> str:
    valor = (valor or '').strip().lower()
    if valor and '@' not in valor:
        valor = f'{valor}@app.com'
    return valor


def verificar_senha(senha_digitada: str, senha_hash_salva: str) -> bool:
    senha_hash_salva = senha_hash_salva or ''
    try:
        identify_hasher(senha_hash_salva)
        return check_password(senha_digitada or '', senha_hash_salva)
    except Exception:
        # Compatibilidade com o hash SHA256 do aplicativo antigo.
        return hmac.compare_digest(hash_senha_antiga(senha_digitada), senha_hash_salva)


def get_client_ip(request) -> str:
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def calcular_dias_restantes(data_expiracao) -> int:
    if not data_expiracao:
        return 0
    if isinstance(data_expiracao, datetime):
        data_expiracao = data_expiracao.date()
    hoje = date.today()
    return max(0, (data_expiracao - hoje).days)


def conta_esta_ativa(conta: Conta) -> bool:
    return bool(conta and int(conta.ativo or 0) == 1)


def conta_trial_expirado(conta: Conta) -> bool:
    return not conta_esta_ativa(conta) and calcular_dias_restantes(conta.trial_expira_em) <= 0


def limite_produtos(conta: Conta) -> int:
    return settings.PRO_LIMITE_PRODUTOS if conta_esta_ativa(conta) else settings.TRIAL_LIMITE_PRODUTOS


def registrar_log(email: str, acao: str) -> None:
    try:
        with connection.cursor() as cursor:
            cursor.execute('INSERT INTO logs (email, acao, criado_em) VALUES (%s, %s, %s)', [email, acao, datetime.now()])
    except Exception:
        pass




def _parse_dt(valor):
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    texto = str(valor).split('.')[0]
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(texto, fmt)
        except Exception:
            pass
    return None


def _codigo_hash(email: str, codigo: str) -> str:
    base = f"{normalizar_email(email)}:{(codigo or '').strip()}:{settings.SECRET_KEY}"
    return hashlib.sha256(base.encode('utf-8')).hexdigest()


def gerar_codigo_numerico() -> str:
    return str(100000 + secrets.randbelow(900000))


def whatsapp_authorization_link(email: str = '') -> str:
    numero = re.sub(r'\D+', '', getattr(settings, 'CADASTRO_AUTORIZACAO_WHATSAPP', '') or '')
    if not numero:
        return ''
    if not numero.startswith('55') and len(numero) in (10, 11):
        numero = '55' + numero
    msg = (
        'Olá, solicitei um código de autorização para criar conta no ValiControl Web.'
        + (f' Usuário/e-mail: {normalizar_email(email)}.' if email else '')
    )
    return f"https://wa.me/{numero}?text={urllib.parse.quote(msg)}"


def _lista_emails_autorizacao() -> list[str]:
    # O destino do código é fixo/travado para não confundir com o remetente da Brevo.
    # Por padrão, o código vai para thiago01268230@gmail.com.
    valor = (
        getattr(settings, 'CADASTRO_EMAIL_TRAVADO', '')
        or getattr(settings, 'CADASTRO_DESTINATARIO_CODIGO', '')
        or getattr(settings, 'CADASTRO_AUTORIZACAO_EMAIL', '')
        or 'thiago01268230@gmail.com'
    )
    emails = [item.strip() for item in re.split(r'[,;\s]+', str(valor)) if item.strip()]
    return emails


def _enviar_email_brevo(destinatarios: list[str], assunto: str, texto: str, html: str = ''):
    api_key = (getattr(settings, 'BREVO_API_KEY', '') or '').strip().strip('"').strip("'")
    sender_email = (getattr(settings, 'BREVO_SENDER_EMAIL', '') or '').strip()
    sender_name = (getattr(settings, 'BREVO_SENDER_NAME', '') or 'ValiControl Web').strip()
    if not api_key or not sender_email:
        return False, 'Configure BREVO_API_KEY e BREVO_SENDER_EMAIL no RunSite.'

    payload = {
        'sender': {'name': sender_name, 'email': sender_email},
        'to': [{'email': email} for email in destinatarios],
        'subject': assunto,
        'textContent': texto,
        'htmlContent': html or texto.replace('\n', '<br>'),
    }
    try:
        response = requests.post(
            'https://api.brevo.com/v3/smtp/email',
            headers={'api-key': api_key, 'accept': 'application/json', 'content-type': 'application/json'},
            json=payload,
            timeout=getattr(settings, 'EMAIL_TIMEOUT', 20),
        )
        if response.status_code >= 400:
            detalhes = response.text[:500]
            return False, f'Brevo recusou o envio ({response.status_code}): {detalhes}'
        return True, None
    except Exception as exc:
        return False, f'Erro ao conectar na Brevo: {exc}'


def solicitar_codigo_autorizacao(email: str, ip: str = ''):
    # Gera um código e envia para o e-mail administrativo configurado.
    # O código nunca é enviado ao usuário solicitante; o administrador repassa se autorizar.
    email = normalizar_email(email)
    if not email:
        return None, 'Informe o usuário/e-mail para solicitar o código.'
    if Conta.objects.filter(email=email).exists():
        return None, 'Este usuário já existe. Use a tela de login.'

    destinatarios = _lista_emails_autorizacao()
    if not destinatarios:
        return None, 'Configure CADASTRO_EMAIL_TRAVADO no RunSite.'

    codigo = gerar_codigo_numerico()
    agora = datetime.now()
    expira = agora + timedelta(minutes=getattr(settings, 'CADASTRO_CODIGO_EXPIRA_MINUTOS', 30))
    codigo_hash = _codigo_hash(email, codigo)

    with connection.cursor() as cursor:
        cursor.execute(
            '''
            INSERT INTO codigos_autorizacao (email, codigo_hash, ip, canal, criado_em, expira_em, usado_em)
            VALUES (%s, %s, %s, %s, %s, %s, NULL)
            ''',
            [email, codigo_hash, ip, 'email_admin', agora, expira],
        )

    assunto = f'Código de autorização ValiControl: {codigo}'
    texto = (
        'Código de autorização para criação de conta no ValiControl Web\n\n'
        f'Código: {codigo}\n'
        f'Conta solicitada: {email}\n'
        f'IP: {ip or "não identificado"}\n'
        f'Validade: {getattr(settings, "CADASTRO_CODIGO_EXPIRA_MINUTOS", 30)} minutos\n\n'
        'Repasse este código somente se você reconhecer e autorizar a criação da conta.'
    )
    html = f'''
    <div style="font-family:Arial,sans-serif;line-height:1.5;color:#0f172a">
      <h2>Código de autorização ValiControl</h2>
      <p>Uma nova conta solicitou autorização para cadastro.</p>
      <p style="font-size:28px;font-weight:800;letter-spacing:4px;background:#f1f5f9;padding:14px;border-radius:10px;display:inline-block">{codigo}</p>
      <p><strong>Conta solicitada:</strong> {email}</p>
      <p><strong>IP:</strong> {ip or 'não identificado'}</p>
      <p><strong>Validade:</strong> {getattr(settings, 'CADASTRO_CODIGO_EXPIRA_MINUTOS', 30)} minutos</p>
      <p>Repasse este código somente se você reconhecer e autorizar a criação da conta.</p>
    </div>
    '''
    ok, erro = _enviar_email_brevo(destinatarios, assunto, texto, html)
    if not ok:
        return None, erro
    registrar_log(email, 'codigo_cadastro_enviado_admin')
    return {'email': email, 'expira_em': expira, 'whatsapp_link': whatsapp_authorization_link(email)}, None


def validar_codigo_autorizacao(email: str, codigo: str):
    # Nesta versão, o cadastro é sempre protegido por código de autorização.
    email = normalizar_email(email)
    codigo = (codigo or '').strip()
    if not codigo:
        return None, 'Informe o código de autorização.'
    if not re.fullmatch(r'\d{6}', codigo):
        return None, 'O código de autorização precisa ter 6 números.'

    esperado = _codigo_hash(email, codigo)
    with connection.cursor() as cursor:
        cursor.execute(
            '''
            SELECT id, expira_em, usado_em
            FROM codigos_autorizacao
            WHERE email=%s AND codigo_hash=%s
            ORDER BY criado_em DESC
            LIMIT 1
            ''',
            [email, esperado],
        )
        row = cursor.fetchone()

    if not row:
        return None, 'Código de autorização inválido.'
    codigo_id, expira_em, usado_em = row
    if usado_em:
        return None, 'Este código já foi usado. Solicite outro código.'
    expira_dt = _parse_dt(expira_em)
    if expira_dt and expira_dt < datetime.now():
        return None, 'Este código expirou. Solicite outro código.'
    return {'id': codigo_id}, None


def marcar_codigo_autorizacao_usado(codigo_id) -> None:
    if not codigo_id:
        return
    with connection.cursor() as cursor:
        cursor.execute('UPDATE codigos_autorizacao SET usado_em=%s WHERE id=%s', [datetime.now(), codigo_id])


def criar_conta(email: str, senha: str, ip: str):
    email = normalizar_email(email)
    if not email or not senha:
        return None, 'Preencha usuário/e-mail e senha.'
    if len(senha) < 4:
        return None, 'A senha precisa ter pelo menos 4 caracteres.'

    if Conta.objects.filter(email=email).exists():
        return None, 'Este usuário já existe.'

    total_ip = Conta.objects.filter(ip=ip, ativo=0).count() if ip else 0
    if total_ip >= settings.TRIAL_MAX_CONTAS_POR_IP:
        return None, 'Limite de contas trial atingido neste IP.'

    agora = datetime.now()
    conta = Conta(
        email=email,
        senha=hash_senha(senha),
        token=gerar_token(),
        criado_em=agora,
        trial_expira_em=agora + timedelta(days=settings.TRIAL_DIAS),
        ativo=0,
        device_id='web',
        plano='trial',
        ip=ip,
    )
    conta.save(force_insert=True)
    registrar_log(email, 'registro_web')
    return conta, None


def autenticar(email: str, senha: str):
    email = normalizar_email(email)
    try:
        conta = Conta.objects.get(email=email)
    except Conta.DoesNotExist:
        return None, 'Usuário não encontrado.'

    if not verificar_senha(senha, conta.senha):
        return None, 'Senha inválida.'

    # Mesmo com trial vencido, deixa entrar para conseguir ver painel e pagar.
    conta.token = gerar_token()
    conta.save(update_fields=['token'])
    registrar_log(email, 'login_web')
    return conta, None


def get_conta_por_email(email: str):
    if not email:
        return None
    try:
        return Conta.objects.get(email=email)
    except Conta.DoesNotExist:
        return None


def produto_status(validade: str) -> str:
    try:
        data = datetime.strptime(str(validade), '%Y-%m-%d').date()
    except Exception:
        return 'sem-data'

    hoje = date.today()
    if data < hoje:
        return 'vencido'
    if (data - hoje).days <= settings.VENCIMENTO_PROXIMO_DIAS:
        return 'proximo'
    return 'ok'


def status_label(status: str) -> str:
    return {
        'vencido': 'Vencido',
        'proximo': 'Próximo',
        'ok': 'OK',
        'sem-data': 'Sem data',
    }.get(status, status)


def formatar_data_br(validade: str) -> str:
    try:
        return datetime.strptime(str(validade), '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return validade or '-'


def produtos_do_usuario(email: str):
    return Produto.objects.filter(user_email=email).order_by('validade', 'nome')


def estatisticas(conta: Conta) -> dict:
    # Lê só a coluna de validade em vez de carregar todos os produtos completos.
    # Isso deixa dashboard e produtos rápidos mesmo com muitos registros.
    qs = Produto.objects.filter(user_email=conta.email).values_list('validade', flat=True)
    total = 0
    vencidos = 0
    proximos = 0
    ok = 0
    for validade in qs.iterator(chunk_size=1000):
        total += 1
        status = produto_status(validade)
        if status == 'vencido':
            vencidos += 1
        elif status == 'proximo':
            proximos += 1
        elif status == 'ok':
            ok += 1
    limite = limite_produtos(conta)
    uso = min(100, int((total / max(1, limite)) * 100))
    expirado = conta_trial_expirado(conta)
    return {
        'total': total,
        'vencidos': vencidos,
        'proximos': proximos,
        'ok': ok,
        'trial_restante': calcular_dias_restantes(conta.trial_expira_em),
        'limite': limite,
        'uso': uso,
        'plano': 'PRO' if conta_esta_ativa(conta) else 'TRIAL',
        'ativo': conta_esta_ativa(conta),
        'trial_expirado': expirado,
        'bloqueado': total >= limite or expirado,
    }


def pode_adicionar_produto(conta: Conta):
    if conta_trial_expirado(conta):
        return False, 'Trial expirado. Atualize para o PRO para continuar cadastrando produtos.'
    total = Produto.objects.filter(user_email=conta.email).count()
    limite = limite_produtos(conta)
    if total >= limite:
        return False, f'Limite atingido: {total}/{limite} produtos.'
    return True, None



def _normalizar_texto_catalogo(valor) -> str:
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = texto.replace('_', ' ').replace('-', ' ')
    return ' '.join(texto.split())


def _normalizar_codigo_catalogo(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip().replace('\u00a0', '').replace(' ', '')
    if texto.endswith('.0') and texto[:-2].isdigit():
        texto = texto[:-2]
    return texto.strip()


CATALOGO_INDEX_VERSION = 'sqlite-stream-v5'
_XLSX_NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
_CELL_REF_RE = re.compile(r'([A-Z]+)')


def _catalogo_db_path() -> Path:
    configured = getattr(settings, 'DATA_SQLITE_PATH', None)
    if configured:
        path = Path(configured)
    else:
        path = Path(settings.DATA_EXCEL_PATH).with_suffix('.sqlite3')

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        return path
    except Exception:
        nome = 'valicontrol_catalogo_' + hashlib.sha1(str(settings.DATA_EXCEL_PATH).encode('utf-8')).hexdigest()[:12] + '.sqlite3'
        return Path(tempfile.gettempdir()) / nome


def _sqlite_connect(path: Path, readonly: bool = False):
    if readonly:
        return sqlite3.connect(f'file:{path}?mode=ro', uri=True)
    return sqlite3.connect(path)


def _arquivo_sha1(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open('rb') as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _excel_signature(path: Path) -> dict:
    stat = path.stat()
    return {
        'version': CATALOGO_INDEX_VERSION,
        'xlsx_size': str(stat.st_size),
        'xlsx_sha1': _arquivo_sha1(path),
    }


def catalogo_index_pronto() -> bool:
    """Verificação rápida do índice do catálogo.

    A versão anterior calculava SHA1 do dados.xlsx em toda consulta. No RunSite
    isso deixava o preenchimento automático lento, porque cada código digitado
    lia o arquivo Excel novamente. Agora a requisição web só confere se o
    SQLite do catálogo existe e tem dados. A validação completa continua sendo
    feita quando o comando indexar_catalogo é executado manualmente.
    """
    db_path = _catalogo_db_path()
    if not db_path.exists():
        return False
    try:
        with _sqlite_connect(db_path, readonly=True) as conn:
            row = conn.execute("SELECT valor FROM catalogo_meta WHERE chave='total_produtos'").fetchone()
            if not row or int(row[0] or 0) <= 0:
                return False
            conn.execute('SELECT 1 FROM catalogo_lookup LIMIT 1').fetchone()
            return True
    except Exception:
        return False


def _xlsx_shared_strings(zf: ZipFile) -> list[str]:
    if 'xl/sharedStrings.xml' not in zf.namelist():
        return []
    strings = []
    with zf.open('xl/sharedStrings.xml') as fp:
        for event, elem in iterparse(fp, events=('end',)):
            if elem.tag == _XLSX_NS + 'si':
                strings.append(''.join((node.text or '') for node in elem.iter(_XLSX_NS + 't')))
                elem.clear()
    return strings


def _xlsx_first_sheet_path(zf: ZipFile) -> str:
    names = zf.namelist()
    if 'xl/worksheets/sheet1.xml' in names:
        return 'xl/worksheets/sheet1.xml'
    for name in names:
        if name.startswith('xl/worksheets/') and name.endswith('.xml'):
            return name
    raise FileNotFoundError('Nenhuma planilha encontrada dentro do XLSX.')


def _cell_col_index(cell_ref: str, fallback: int) -> int:
    match = _CELL_REF_RE.match(cell_ref or '')
    if not match:
        return fallback
    total = 0
    for ch in match.group(1):
        total = total * 26 + (ord(ch) - ord('A') + 1)
    return total - 1


def _xlsx_cell_value(cell, shared_strings: list[str]) -> str:
    tipo = cell.attrib.get('t')
    if tipo == 'inlineStr':
        inline = cell.find(_XLSX_NS + 'is')
        if inline is None:
            return ''
        return ''.join((node.text or '') for node in inline.iter(_XLSX_NS + 't'))

    value = cell.find(_XLSX_NS + 'v')
    if value is None or value.text is None:
        return ''
    raw = value.text
    if tipo == 's':
        try:
            return shared_strings[int(raw)]
        except Exception:
            return ''
    return raw


def _iter_xlsx_rows(path: Path):
    """Lê o XLSX por XML, sem carregar o arquivo inteiro na memória.

    A planilha do sistema antigo declara a dimensão como A1, por isso o modo
    streaming do openpyxl ignora as demais colunas. A leitura direta do XML
    evita esse problema e é leve o suficiente para o RunSite.
    """
    with ZipFile(path) as zf:
        shared_strings = _xlsx_shared_strings(zf)
        sheet_path = _xlsx_first_sheet_path(zf)
        with zf.open(sheet_path) as fp:
            for event, row in iterparse(fp, events=('end',)):
                if row.tag != _XLSX_NS + 'row':
                    continue
                values_by_col = {}
                max_col = -1
                fallback_col = 0
                for cell in row.findall(_XLSX_NS + 'c'):
                    col_idx = _cell_col_index(cell.attrib.get('r', ''), fallback_col)
                    fallback_col = col_idx + 1
                    values_by_col[col_idx] = _xlsx_cell_value(cell, shared_strings)
                    max_col = max(max_col, col_idx)
                yield [values_by_col.get(i, '') for i in range(max_col + 1)] if max_col >= 0 else []
                row.clear()


def _encontrar_coluna(headers: list[str], *nomes: str):
    alvos = [_normalizar_texto_catalogo(nome) for nome in nomes]
    for alvo in alvos:
        if alvo in headers:
            return headers.index(alvo)
    return None


def indexar_catalogo_produtos(force: bool = False) -> int:
    """Cria um índice SQLite leve para consulta por código/GTIN.

    Retorna a quantidade de produtos indexados. Não apaga nem altera os produtos
    cadastrados pelos usuários; o banco gerado é apenas um cache da base
    data/dados.xlsx.
    """
    xlsx_path = Path(settings.DATA_EXCEL_PATH)
    if not xlsx_path.exists():
        logger.warning('Catálogo XLSX não encontrado: %s', xlsx_path)
        return 0

    db_path = _catalogo_db_path()
    if not force and catalogo_index_pronto():
        return CatalogoProdutos(db_path).total_produtos()

    temp_path = db_path.with_suffix('.tmp.sqlite3')
    try:
        if temp_path.exists():
            temp_path.unlink()
    except Exception:
        pass

    total = 0
    header_encontrado = False
    idx_produto = idx_descricao = idx_gtin = None
    produtos_batch = []
    lookup_batch = []

    conn = _sqlite_connect(temp_path)
    try:
        conn.execute('PRAGMA journal_mode=OFF')
        conn.execute('PRAGMA synchronous=OFF')
        conn.execute('PRAGMA temp_store=MEMORY')
        conn.execute('CREATE TABLE catalogo_meta (chave TEXT PRIMARY KEY, valor TEXT)')
        conn.execute('CREATE TABLE catalogo_produtos (id INTEGER PRIMARY KEY, codigo TEXT, nome TEXT NOT NULL, gtin TEXT)')
        conn.execute('CREATE TABLE catalogo_lookup (chave TEXT PRIMARY KEY, produto_id INTEGER NOT NULL)')

        def flush():
            if produtos_batch:
                conn.executemany(
                    'INSERT INTO catalogo_produtos (id, codigo, nome, gtin) VALUES (?, ?, ?, ?)',
                    produtos_batch,
                )
                produtos_batch.clear()
            if lookup_batch:
                conn.executemany(
                    'INSERT OR REPLACE INTO catalogo_lookup (chave, produto_id) VALUES (?, ?)',
                    lookup_batch,
                )
                lookup_batch.clear()

        for row in _iter_xlsx_rows(xlsx_path):
            if not header_encontrado:
                normalizados = [_normalizar_texto_catalogo(c) for c in row]
                if 'produto' in normalizados and ('descricao' in normalizados or 'nome' in normalizados):
                    idx_produto = _encontrar_coluna(normalizados, 'Produto', 'Código', 'Codigo', 'Código interno', 'Codigo interno')
                    idx_descricao = _encontrar_coluna(normalizados, 'Descrição', 'Descricao', 'Nome', 'Nome do produto', 'Produto descrição')
                    idx_gtin = _encontrar_coluna(normalizados, 'GTIN Principal', 'GTIN', 'EAN', 'Código de barras', 'Codigo de barras')
                    header_encontrado = idx_descricao is not None and (idx_produto is not None or idx_gtin is not None)
                continue

            codigo = row[idx_produto] if idx_produto is not None and idx_produto < len(row) else ''
            nome = row[idx_descricao] if idx_descricao is not None and idx_descricao < len(row) else ''
            gtin = row[idx_gtin] if idx_gtin is not None and idx_gtin < len(row) else ''

            nome_s = str(nome or '').strip()
            codigo_s = _normalizar_codigo_catalogo(codigo)
            gtin_s = _normalizar_codigo_catalogo(gtin)

            if not nome_s or (not codigo_s and not gtin_s):
                continue

            total += 1
            produtos_batch.append((total, codigo_s or gtin_s, nome_s, gtin_s))
            if codigo_s and codigo_s.lower() != 'none':
                lookup_batch.append((codigo_s, total))
            if gtin_s and gtin_s.lower() != 'none':
                lookup_batch.append((gtin_s, total))

            if total % 2000 == 0:
                flush()

        flush()
        for chave, valor in _excel_signature(xlsx_path).items():
            conn.execute('INSERT INTO catalogo_meta (chave, valor) VALUES (?, ?)', (chave, valor))
        conn.execute('INSERT INTO catalogo_meta (chave, valor) VALUES (?, ?)', ('total_produtos', str(total)))
        conn.commit()
    except Exception:
        conn.rollback()
        logger.exception('Falha ao indexar catálogo de produtos em %s', xlsx_path)
        raise
    finally:
        conn.close()

    os.replace(temp_path, db_path)
    return total


class CatalogoProdutos:
    """Objeto parecido com dict, mas consultado em SQLite para não estourar memória."""

    def __init__(self, db_path: Path | None = None):
        self.db_path = Path(db_path or _catalogo_db_path())

    def _ensure_ready(self):
        # Não indexa automaticamente durante uma requisição web.
        # Isso evita que o RunSite derrube o site por memória/tempo caso o
        # arquivo data/dados.sqlite3 não tenha sido enviado ao GitHub.
        if not catalogo_index_pronto():
            raise FileNotFoundError(
                'Catálogo ainda não indexado. Envie data/dados.sqlite3 ou rode: python manage.py indexar_catalogo --force'
            )

    def _fetchone(self, sql: str, params=()):
        self._ensure_ready()
        with _sqlite_connect(self.db_path, readonly=True) as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute(sql, params).fetchone()

    def total_produtos(self) -> int:
        try:
            row = self._fetchone("SELECT valor FROM catalogo_meta WHERE chave='total_produtos'")
            return int(row['valor']) if row else 0
        except Exception:
            return 0

    def __len__(self) -> int:
        return self.total_produtos()

    def get(self, codigo: str, default=None):
        codigo = _normalizar_codigo_catalogo(codigo)
        if not codigo:
            return default
        try:
            row = self._fetchone(
                """
                SELECT p.codigo, p.nome, p.gtin
                FROM catalogo_lookup l
                JOIN catalogo_produtos p ON p.id = l.produto_id
                WHERE l.chave = ?
                LIMIT 1
                """,
                (codigo,),
            )
            if not row:
                return default
            return {'codigo': row['codigo'] or '', 'nome': row['nome'] or '', 'gtin': row['gtin'] or ''}
        except Exception:
            logger.exception('Falha ao consultar catálogo pelo código %s', codigo)
            return default

    def __contains__(self, codigo: str) -> bool:
        return self.get(codigo) is not None

    def __getitem__(self, codigo: str):
        item = self.get(codigo)
        if item is None:
            raise KeyError(codigo)
        return item

    def items(self):
        self._ensure_ready()
        with _sqlite_connect(self.db_path, readonly=True) as conn:
            conn.row_factory = sqlite3.Row
            for row in conn.execute('SELECT codigo, nome, gtin FROM catalogo_produtos ORDER BY id'):
                chave = row['gtin'] or row['codigo'] or ''
                yield chave, {'codigo': row['codigo'] or '', 'nome': row['nome'] or '', 'gtin': row['gtin'] or ''}


@lru_cache(maxsize=1)
def carregar_catalogo_produtos():
    return CatalogoProdutos()


@lru_cache(maxsize=4096)
def buscar_catalogo(codigo: str):
    # Cache pequeno para deixar leituras repetidas instantâneas, sem carregar
    # todo o catálogo na memória.
    return carregar_catalogo_produtos().get(codigo)

def limpar_validade(valor) -> str:
    """Aceita yyyy-mm-dd, dd/mm/yyyy ou data Excel e devolve yyyy-mm-dd."""
    if not valor:
        return ''
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    if isinstance(valor, date):
        return valor.strftime('%Y-%m-%d')
    texto = str(valor).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(texto, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return texto


def montar_produto_dict(produto: Produto) -> dict:
    status = produto_status(produto.validade)
    return {
        'id': produto.id,
        'codigo': produto.codigo or '',
        'nome': produto.nome or '',
        'validade_original': produto.validade or '',
        'validade': formatar_data_br(produto.validade),
        'quantidade': produto.quantidade,
        'tipo_qtd': produto.tipo_qtd or 'Un',
        'status': status,
        'status_label': status_label(status),
    }


def exportar_produtos_xlsx(conta: Conta) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produtos'
    ws.append(['ID', 'Código', 'Nome', 'Validade', 'Quantidade', 'Tipo', 'Status'])
    for produto in produtos_do_usuario(conta.email):
        item = montar_produto_dict(produto)
        ws.append([
            item['id'], item['codigo'], item['nome'], item['validade'],
            item['quantidade'], item['tipo_qtd'], item['status_label']
        ])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 48)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def exportar_produtos_csv(conta: Conta) -> str:
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['ID', 'Código', 'Nome', 'Validade', 'Quantidade', 'Tipo', 'Status'])
    for produto in produtos_do_usuario(conta.email):
        item = montar_produto_dict(produto)
        writer.writerow([
            item['id'], item['codigo'], item['nome'], item['validade'],
            item['quantidade'], item['tipo_qtd'], item['status_label']
        ])
    return output.getvalue()


def importar_produtos_de_planilha(conta: Conta, arquivo) -> tuple[int, list[str]]:
    """Importa XLSX/CSV com colunas código, nome, validade, quantidade e tipo."""
    erros = []
    criados = 0
    nome_arquivo = (getattr(arquivo, 'name', '') or '').lower()

    def normalizar_header(h):
        return str(h or '').strip().lower().replace('ó', 'o').replace('ç', 'c')

    def processar_linhas(linhas):
        nonlocal criados
        linhas = list(linhas)
        if not linhas:
            return
        headers = [normalizar_header(h) for h in linhas[0]]

        def idx(*nomes):
            for nome in nomes:
                nome = normalizar_header(nome)
                if nome in headers:
                    return headers.index(nome)
            return None

        idx_codigo = idx('codigo', 'código', 'gtin', 'codigo/gtin')
        idx_nome = idx('nome', 'descricao', 'descrição', 'produto')
        idx_validade = idx('validade', 'vencimento', 'data')
        idx_qtd = idx('quantidade', 'qtd')
        idx_tipo = idx('tipo', 'tipo_qtd', 'unidade')

        if idx_codigo is None or idx_nome is None or idx_validade is None:
            erros.append('A planilha precisa ter pelo menos: código, nome e validade.')
            return

        for numero, row in enumerate(linhas[1:], start=2):
            try:
                codigo = str(row[idx_codigo] or '').replace('.0', '').strip() if idx_codigo < len(row) else ''
                nome = str(row[idx_nome] or '').strip() if idx_nome < len(row) else ''
                validade = limpar_validade(row[idx_validade] if idx_validade < len(row) else '')
                qtd_raw = row[idx_qtd] if idx_qtd is not None and idx_qtd < len(row) else 0
                tipo = str(row[idx_tipo] or 'Un').strip() if idx_tipo is not None and idx_tipo < len(row) else 'Un'
                quantidade = int(float(qtd_raw or 0))
                datetime.strptime(validade, '%Y-%m-%d')
                if not codigo or not nome:
                    continue
                ok, erro = pode_adicionar_produto(conta)
                if not ok:
                    erros.append(erro)
                    break
                Produto.objects.create(
                    codigo=codigo,
                    nome=nome,
                    validade=validade,
                    quantidade=max(0, quantidade),
                    tipo_qtd=tipo or 'Un',
                    user_email=conta.email,
                )
                criados += 1
            except Exception:
                erros.append(f'Linha {numero}: dados inválidos.')

    if nome_arquivo.endswith('.csv'):
        conteudo = arquivo.read().decode('utf-8-sig', errors='ignore')
        sample = conteudo[:2048]
        delimiter = ';' if sample.count(';') >= sample.count(',') else ','
        processar_linhas(csv.reader(StringIO(conteudo), delimiter=delimiter))
    else:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        processar_linhas(ws.iter_rows(values_only=True))

    if criados:
        registrar_log(conta.email, f'importacao_{criados}_produtos_web')
    return criados, erros[:5]


ASAAS_BASE_URL = getattr(settings, 'ASAAS_BASE_URL', 'https://api.asaas.com/v3')


def asaas_headers():
    if not settings.ASAAS_API_KEY:
        return None
    return {
        'access_token': settings.ASAAS_API_KEY,
        'Content-Type': 'application/json',
    }


def criar_cliente_asaas(email: str):
    headers = asaas_headers()
    if not headers:
        return None, 'ASAAS_API_KEY não configurada.'

    try:
        response = requests.post(
            f'{ASAAS_BASE_URL}/customers',
            headers=headers,
            json={'name': email, 'email': email},
            timeout=30,
        )
        try:
            data = response.json()
        except Exception:
            data = {}
        if response.status_code >= 400 or 'errors' in data:
            consulta = requests.get(
                f'{ASAAS_BASE_URL}/customers',
                headers=headers,
                params={'email': email},
                timeout=30,
            )
            clientes = consulta.json().get('data', [])
            if clientes:
                return clientes[0].get('id'), None
            return None, 'Não foi possível criar/consultar cliente no Asaas.'
        return data.get('id'), None
    except Exception as exc:
        return None, f'Falha ao conectar no Asaas: {exc}'


def criar_pagamento_pix(email: str):
    headers = asaas_headers()
    if not headers:
        return None, 'ASAAS_API_KEY não configurada no ambiente.'

    customer_id, erro = criar_cliente_asaas(email)
    if erro:
        return None, erro

    try:
        response = requests.post(
            f'{ASAAS_BASE_URL}/payments',
            headers=headers,
            json={
                'customer': customer_id,
                'billingType': 'PIX',
                'value': settings.PIX_VALOR,
                'dueDate': (date.today() + timedelta(days=1)).isoformat(),
                'description': settings.PAGAMENTO_DESCRICAO,
                'externalReference': email,
            },
            timeout=30,
        )
        try:
            data = response.json()
        except Exception:
            data = {}
        if response.status_code >= 400 or 'errors' in data:
            return None, 'Erro ao criar pagamento no Asaas. Confira a ASAAS_API_KEY.'

        payment_id = data.get('id')
        qr_response = requests.get(
            f'{ASAAS_BASE_URL}/payments/{payment_id}/pixQrCode',
            headers=headers,
            timeout=30,
        )
        qr_data = qr_response.json()
        if qr_response.status_code >= 400 or 'errors' in qr_data:
            return None, 'Erro ao gerar QR Code PIX.'

        with connection.cursor() as cursor:
            cursor.execute(
                'INSERT INTO pagamentos (payment_id, email, status, criado_em) VALUES (%s, %s, %s, %s) ON CONFLICT (payment_id) DO NOTHING'
                if connection.vendor == 'postgresql'
                else 'INSERT OR IGNORE INTO pagamentos (payment_id, email, status, criado_em) VALUES (%s, %s, %s, %s)',
                [payment_id, email, data.get('status', 'PENDING'), datetime.now()],
            )

        return {
            'payment_id': payment_id,
            'qr': qr_data.get('payload'),
            'qr_base64': qr_data.get('encodedImage'),
            'valor': settings.PIX_VALOR,
        }, None
    except Exception as exc:
        return None, f'Falha ao criar PIX: {exc}'


def ativar_usuario(email: str, dias: int = 30) -> bool:
    try:
        conta = Conta.objects.get(email=normalizar_email(email))
    except Conta.DoesNotExist:
        return False
    conta.ativo = 1
    conta.plano = 'pago'
    conta.trial_expira_em = datetime.now() + timedelta(days=dias)
    conta.save(update_fields=['ativo', 'plano', 'trial_expira_em'])
    registrar_log(conta.email, 'pagamento_aprovado_web')
    return True
